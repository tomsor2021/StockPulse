# -*- coding: utf-8 -*-
"""清空 stock_basic 表并从 Excel 文件重新导入数据"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from database.db import get_connection
from datetime import datetime

EXCEL_PATH = r"F:\mywork\claw工作\财经\股票基金代码表.xlsx"

SHEET_MAP = {
    "A股股票": {"market_col": "市场", "default_market": None},  # use column value
    "港股股票": {"market_col": "市场", "default_market": None},
    "A股基金": {"market_col": None, "default_market": "基金"},
}


def read_sheet_data(ws, config):
    """读取一个 sheet 的数据，返回 (code, name, market) 列表"""
    rows = []
    market_col = config["market_col"]
    default_market = config["default_market"]

    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            # find column indices
            try:
                code_idx = header.index("代码")
                name_idx = header.index("名称")
                mkt_idx = header.index(market_col) if market_col else None
            except ValueError:
                print(f"  ❌ 表头缺少必要列: {header}")
                return rows
            continue
        if row[code_idx] is None:
            continue
        code = str(row[code_idx]).strip().split(".")[0]
        if not code:
            continue
        name = str(row[name_idx]).strip()
        if market_col and mkt_idx is not None:
            market = str(row[mkt_idx]).strip()
        else:
            market = default_market
        rows.append((code, name, market))
    return rows


def main():
    print("=" * 60)
    print("导入股票基金代码表到 stock_basic")
    print("=" * 60)

    # 1. 读取 Excel
    print(f"\n📂 读取 Excel 文件: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

    all_data = []
    sheet_counts = {}
    for sheet_name, config in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet '{sheet_name}' 不存在，跳过")
            continue
        ws = wb[sheet_name]
        data = read_sheet_data(ws, config)
        sheet_counts[sheet_name] = len(data)
        all_data.extend(data)
        print(f"  ✅ {sheet_name}: {len(data)} 条记录")

    wb.close()
    print(f"\n📊 总计读取 {len(all_data)} 条记录")
    if not all_data:
        print("❌ 没有数据可导入")
        return

    # 2. 清空表并导入
    print(f"\n🗄️  连接数据库并导入...")
    conn = get_connection()

    # 清空
    conn.execute("DELETE FROM stock_basic")
    conn.commit()
    print("  ✅ stock_basic 表已清空")

    # 批量插入
    count = 0
    now = datetime.now()
    for code, name, market in all_data:
        try:
            conn.execute(
                "INSERT OR IGNORE INTO stock_basic (code, name, market, updated_at) VALUES (?, ?, ?, ?)",
                (code, name, market, now)
            )
            count += 1
        except Exception as e:
            print(f"  ⚠️  插入失败: {code} {name} - {e}")

    conn.commit()
    print(f"  ✅ 成功导入 {count} 条记录")

    # 3. 验证
    print(f"\n🔍 验证导入结果...")
    cur = conn.execute("SELECT market, COUNT(*) FROM stock_basic GROUP BY market")
    total = 0
    for row in cur:
        print(f"  {row[0]}: {row[1]} 条")
        total += row[1]
    print(f"  ─────────────")
    print(f"  合计: {total} 条")
    conn.close()

    print(f"\n🎉 导入完成！")


if __name__ == "__main__":
    main()
